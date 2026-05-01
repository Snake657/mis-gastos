#!/usr/bin/env python3
"""
Regenera /deuda-publica/data.json desde Finanzas (Min. Economía).

Combina:
  - Serie ANUAL 1992-2018 (sheet A.2.5 del Excel trimestral más reciente,
    fila "I- TOTAL DEUDA PÚBLICA BRUTA").
  - Serie MENSUAL 2019-actualidad (boletín mensual, sheet A.1, fila
    "A- DEUDA BRUTA (I+II+III)").

Se ejecuta cada vez que Finanzas publica un nuevo boletín mensual (~mes y medio
de delay desde el cierre del mes). Requiere `openpyxl` y `requests`.

Uso:
    cd canuto.ar/admin/deuda-publica
    python regenerar.py            # baja archivos, parsea, sobrescribe data.json
    python regenerar.py --dry-run  # solo imprime, no escribe
"""

import sys, json, re, argparse, datetime as dt, urllib.request, urllib.parse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: falta openpyxl. Instalá con: pip install openpyxl")
    sys.exit(1)

# ── Configuración ────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent.parent
OUT  = ROOT / 'deuda-publica' / 'data.json'
TMP  = Path(__file__).resolve().parent / '_tmp'
TMP.mkdir(exist_ok=True)

URL_PAGINA_MENSUAL    = 'https://www.argentina.gob.ar/economia/finanzas/datos-mensuales'
URL_PAGINA_TRIMESTRAL = 'https://www.argentina.gob.ar/economia/finanzas/datos-trimestrales-de-la-deuda'

MESES_ES = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
            'jul':7,'ago':8,'sep':9,'sept':9,'oct':10,'nov':11,'dic':12}

# ── Helpers ──────────────────────────────────────────────────────────────────
def fetch_html(url):
    req = urllib.request.Request(url, headers={'User-Agent': 'canuto.ar/regen 1.0'})
    with urllib.request.urlopen(req, timeout=30) as r:
        return r.read().decode('utf-8', errors='replace')

def descargar(url, dest):
    print(f"  ↓ {url}")
    req = urllib.request.Request(url, headers={'User-Agent': 'canuto.ar/regen 1.0'})
    with urllib.request.urlopen(req, timeout=60) as r, open(dest, 'wb') as f:
        f.write(r.read())
    return dest

def parsear_fecha(h):
    """Parsea headers de Excel: datetime, '31/12/92', 'ene-26 (*)'.
    Devuelve (iso_str, es_preliminar) o (None, False)."""
    if h is None: return None, False
    if isinstance(h, dt.datetime):
        return h.strftime('%Y-%m-%d'), False
    s = str(h)
    prelim = ('(*)' in s) or s.strip().endswith('*')
    s_clean = s.replace('(*)','').replace('*','').strip()
    if '/' in s_clean:
        try:
            d, m, y = s_clean.split('/')
            y = int(y)
            if y < 50: y += 2000
            elif y < 100: y += 1900
            return f"{y}-{int(m):02d}-{int(d):02d}", prelim
        except Exception: pass
    if '-' in s_clean:
        try:
            mes, yy = s_clean.split('-')
            mes = mes.lower().rstrip('.')[:3]
            if mes in MESES_ES:
                yy = int(yy)
                yy = 2000 + yy if yy < 50 else (1900 + yy if yy < 100 else yy)
                return f"{yy}-{MESES_ES[mes]:02d}-01", prelim
        except Exception: pass
    return None, False

def encontrar_url_xlsx(html, patron):
    """Encuentra la primera URL .xlsx en el HTML que matchee el patrón."""
    rx = re.compile(rf'(https?://[^"\'<>]*{patron}[^"\'<>]*\.xlsx)', re.IGNORECASE)
    m = rx.search(html.replace('blank:#', ''))
    return m.group(1) if m else None

# ── Paso 1: descargar boletín mensual más reciente ───────────────────────────
def descargar_mensual():
    print("→ Buscando boletín mensual...")
    html = fetch_html(URL_PAGINA_MENSUAL)
    url = encontrar_url_xlsx(html, r'boletin_mensual')
    if not url:
        raise RuntimeError(f"No se encontró URL del boletín en {URL_PAGINA_MENSUAL}")
    dest = TMP / 'boletin_mensual.xlsx'
    descargar(url, dest)
    return dest, url

# ── Paso 2: descargar Excel trimestral más reciente (para serie histórica) ───
def descargar_trimestral():
    print("→ Buscando Excel trimestral más reciente...")
    html = fetch_html(URL_PAGINA_TRIMESTRAL)
    # El más reciente aparece primero en el HTML
    url = encontrar_url_xlsx(html, r'deuda_publica_\d{2}-\d{2}-\d{4}')
    if not url:
        raise RuntimeError(f"No se encontró URL del Excel trimestral en {URL_PAGINA_TRIMESTRAL}")
    dest = TMP / 'deuda_trimestral.xlsx'
    descargar(url, dest)
    return dest, url

# ── Paso 3: parsear ──────────────────────────────────────────────────────────
def parsear_anual(xlsx_path):
    """Lee sheet A.2.5, fila 17 ('I- TOTAL DEUDA PÚBLICA BRUTA')."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'A.2.5' not in wb.sheetnames:
        raise RuntimeError(f"Sheet A.2.5 no existe en {xlsx_path}. Sheets: {wb.sheetnames}")
    ws = wb['A.2.5']
    header = list(ws.iter_rows(min_row=12, max_row=12, values_only=True))[0]

    # Encontrar la fila correcta — buscar por texto "TOTAL DEUDA PÚBLICA BRUTA" en col B
    fila_total = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        label = row[1] if len(row) > 1 else None
        if label and 'TOTAL DEUDA PÚBLICA BRUTA' in str(label).upper():
            fila_total = i
            break
    if fila_total is None:
        raise RuntimeError("No se encontró fila 'TOTAL DEUDA PÚBLICA BRUTA' en A.2.5")

    valores = list(ws.iter_rows(min_row=fila_total, max_row=fila_total, values_only=True))[0]
    out = []
    for h, v in zip(header[2:], valores[2:]):
        f, _ = parsear_fecha(h)
        if not f or v is None: continue
        out.append({'fecha': f, 'valor': round(float(v), 1), 'fuente': 'anual'})
    out.sort(key=lambda x: x['fecha'])
    print(f"  ✓ {len(out)} puntos anuales: {out[0]['fecha']} → {out[-1]['fecha']}")
    return out

def parsear_mensual(xlsx_path):
    """Lee sheet A.1, fila 'A- DEUDA BRUTA (I+II+III)'."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'A.1' not in wb.sheetnames:
        raise RuntimeError(f"Sheet A.1 no existe en {xlsx_path}. Sheets: {wb.sheetnames}")
    ws = wb['A.1']
    header = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]

    fila_total = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        label = row[1] if len(row) > 1 else None
        if label and re.search(r'^A-\s*DEUDA\s*BRUTA', str(label), re.IGNORECASE):
            fila_total = i
            break
    if fila_total is None:
        raise RuntimeError("No se encontró fila 'A- DEUDA BRUTA' en A.1")

    valores = list(ws.iter_rows(min_row=fila_total, max_row=fila_total, values_only=True))[0]
    out = []
    for h, v in zip(header[2:], valores[2:]):
        f, prelim = parsear_fecha(h)
        if not f or v is None: continue
        item = {'fecha': f, 'valor': round(float(v), 1), 'fuente': 'mensual'}
        if prelim: item['preliminar'] = True
        out.append(item)
    out.sort(key=lambda x: x['fecha'])
    print(f"  ✓ {len(out)} puntos mensuales: {out[0]['fecha']} → {out[-1]['fecha']}")
    return out

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true', help='No escribir data.json')
    args = ap.parse_args()

    xlsx_mens, url_mens = descargar_mensual()
    xlsx_tri,  url_tri  = descargar_trimestral()

    print("\n→ Parseando...")
    anual    = parsear_anual(xlsx_tri)
    mensual  = parsear_mensual(xlsx_mens)

    # Combinar: anual hasta 2018-12-31, mensual desde 2019-01-01
    combinado = [x for x in anual if x['fecha'] < '2019-01-01'] + mensual
    combinado.sort(key=lambda x: x['fecha'])

    salida = {
        'fuente': 'Secretaría de Finanzas - Ministerio de Economía',
        'metodologia': 'Stock de deuda bruta de la Administración Central, en millones de USD. No incluye cupón PBI.',
        'urls_origen': {
            'mensual_2019_actualidad': url_mens,
            'anual_1992_2018': url_tri + ' (sheet A.2.5)',
        },
        'notas': '1992-2018: serie anual al 31/12. 2019 en adelante: mensual. Marca preliminar (*) puede revisarse.',
        'actualizacion': dt.datetime.now().strftime('%Y-%m-%d'),
        'datos': combinado,
    }

    print(f"\n→ Total combinado: {len(combinado)} puntos")
    print(f"  Primero: {combinado[0]['fecha']} = {combinado[0]['valor']:,.1f} M USD")
    print(f"  Último:  {combinado[-1]['fecha']} = {combinado[-1]['valor']:,.1f} M USD"
          f"{' (preliminar)' if combinado[-1].get('preliminar') else ''}")

    if args.dry_run:
        print("\n[--dry-run] No se escribió nada.")
        return

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(salida, f, indent=1, ensure_ascii=False)
    print(f"\n✓ Escrito {OUT} ({OUT.stat().st_size:,} bytes)")

if __name__ == '__main__':
    main()
